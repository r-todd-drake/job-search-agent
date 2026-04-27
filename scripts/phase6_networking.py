# ==============================================
# phase6_networking.py
# Generates warmth-calibrated LinkedIn and email
# outreach messages for professional networking.
#
# Reads contact data from contact_pipeline.xlsx.
# Writes stage advances back on interactive confirm.
# No automated sending -- output is terminal only.
#
# Usage:
#   python -m scripts.phase6_networking \
#     --contact "Jane Smith" --stage 1
#   python -m scripts.phase6_networking \
#     --contact "Jane Smith" --stage 2 --role acme-systems-se
#   python -m scripts.phase6_networking --list
# ==============================================

import os
import sys
import argparse
from datetime import date
from anthropic import Anthropic
from dotenv import load_dotenv
import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from scripts.utils.pii_filter import strip_pii
from scripts.utils import candidate_config
from scripts.config import CONTACTS_TRACKER_PATH, MODEL_SONNET as MODEL

load_dotenv()

# ==============================================
# SYSTEM PROMPT
# ==============================================

SYSTEM_PROMPT = """You are an expert career coach writing professional networking messages \
for senior defense and systems engineering professionals. Your messages are specific, \
genuine, and appropriately brief.

You always:
- Use en dashes, never em dashes
- Write specific over generic
- Match the directness of the message to the warmth of the relationship
- Keep connection requests within their character limits

You never:
- Use hollow openers like "I came across your profile" or "Hope this message finds you well"
- Invent or guess shared history not explicitly provided
- Reference a specific role unless one is provided
- Use filler phrases or overly formal language"""

# ==============================================
# XLSX I/O
# ==============================================

COLUMNS = [
    "contact_name", "company", "title", "linkedin_url", "warmth",
    "source", "first_contact", "response_date", "stage", "status",
    "role_activated", "referral_bonus", "notes",
]


def load_contacts(path: str) -> list:
    """Load all contacts from xlsx. Returns list of dicts keyed by column name.
    datetime values (openpyxl date cells) are normalized to date objects."""
    import datetime as _dt
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    contacts = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        record = dict(zip(headers, row))
        for k, v in record.items():
            if isinstance(v, _dt.datetime):
                record[k] = v.date()
        contacts.append(record)
    return contacts


def find_contact(contacts: list, name: str) -> dict:
    """Case-insensitive partial match on contact_name. Raises ValueError if ambiguous or not found.
    Ambiguous means multiple distinct names match; duplicate rows for the same name return the first."""
    name_lower = name.lower()
    matches = [c for c in contacts if name_lower in (c.get("contact_name") or "").lower()]
    if len(matches) == 0:
        raise ValueError(f"Contact '{name}' not found in contact_pipeline.xlsx.")
    distinct_names = list(dict.fromkeys(c["contact_name"] for c in matches))
    if len(distinct_names) > 1:
        raise ValueError(f"Contact '{name}' is ambiguous -- matched: {', '.join(distinct_names)}. Use a more specific name.")
    return matches[0]


def update_contact(path: str, contact_name: str, updates: dict) -> None:
    """Write field updates back to the xlsx row matching contact_name."""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2):
        if row[headers.index("contact_name")].value == contact_name:
            for field, value in updates.items():
                if field == "response_date":
                    continue  # never written by script
                col_idx = headers.index(field)
                row[col_idx].value = value
            break
    wb.save(path)


# ==============================================
# CANDIDATE CONTEXT
# ==============================================

def _build_candidate_context(candidate: dict) -> str:
    """Extract networking-relevant candidate context from candidate_config dict."""
    lines = [
        f"Name: {os.getenv('CANDIDATE_NAME', '[CANDIDATE]')}",
        f"Location: {os.getenv('CANDIDATE_LOCATION', '[LOCATION]')}",
    ]
    cl = candidate.get("clearance", {})
    if cl:
        lines.append(f"Clearance: {cl.get('status', '')} {cl.get('level', '')}")
    for svc in candidate.get("military", {}).get("service", []):
        branch = svc.get("branch", "")
        dates = svc.get("dates", "")
        if branch:
            lines.append(f"Military: {branch} {dates}".strip())
    prog = candidate.get("confirmed_skills", {}).get("programming", "")
    if prog:
        lines.append(f"Technical skills: {prog}")
    return "\n".join(lines)


# ==============================================
# WARMTH HELPER
# ==============================================

def _warmth_context(warmth: str) -> str:
    """Return placeholder instruction for Claude based on warmth tier."""
    w = warmth.lower()
    if "acquaintance" in w:
        return (
            "Include the placeholder [HOW YOU KNOW THIS PERSON] exactly where you would "
            "reference the shared connection. Do not invent or guess the shared context. "
            "Keep surrounding text to approximately 180 characters to leave room for the fill."
        )
    if "former colleague" in w:
        return (
            "Include the placeholder [WHERE YOU WORKED TOGETHER] exactly where you would "
            "reference the shared employer or project. Do not invent or guess the shared context. "
            "Keep surrounding text to approximately 180 characters to leave room for the fill."
        )
    return ""


# ==============================================
# PROMPT BUILDERS
# ==============================================

def _build_stage1_prompt(contact: dict, candidate: dict) -> str:
    warmth_instruction = _warmth_context(contact.get("warmth", ""))
    char_limit = "180 characters for the surrounding text" if warmth_instruction else "300 characters total"

    parts = [
        f"Write a LinkedIn connection request from {os.getenv('CANDIDATE_NAME', '[CANDIDATE]')} "
        f"to {contact['contact_name']}, a {contact.get('title', 'professional')} at {contact.get('company', 'their company')}.",
        "",
        "Candidate background:",
        _build_candidate_context(candidate),
        "",
        f"Warmth level: {contact.get('warmth', 'Cold')}",
        f"Notes on the relationship: {contact.get('notes') or 'No prior contact.'}",
        "",
        "Connection request requirements:",
        f"- HARD LIMIT: {char_limit}",
        "- Specific and genuine -- no 'I came across your profile' openers",
        "- No role ask -- relationship-building only",
        "- Write the connection request text only -- no labels, no preamble",
    ]

    if warmth_instruction:
        parts.append(f"- {warmth_instruction}")

    parts.extend([
        "",
        "Then write a follow-up message for if the candidate is already connected with "
        f"{contact['contact_name']}. The follow-up should be 2-3 sentences -- concise and warm.",
        "Separate the two outputs with exactly this delimiter on its own line: ---FOLLOW-UP---",
        "Format: [connection request text]\\n---FOLLOW-UP---\\n[follow-up message text]",
    ])

    return "\n".join(parts)


def _build_stage2_prompt(contact: dict, candidate: dict, jd_text: str) -> str:
    warmth = contact.get("warmth", "Cold")
    warmth_instruction = _warmth_context(warmth)
    referral_bonus = contact.get("referral_bonus")

    directness = {
        "cold": "professional and neutral -- you have no prior relationship",
        "acquaintance": "warm but measured -- you have a limited prior connection",
        "former colleague": "collegial and direct -- you have a real working history together",
        "strong": "direct and personal -- this is a close professional contact",
    }.get(warmth.lower(), "professional and warm")

    parts = [
        f"Write a LinkedIn message or email from {os.getenv('CANDIDATE_NAME', '[CANDIDATE]')} "
        f"to {contact['contact_name']}, a {contact.get('title', 'professional')} at {contact.get('company', 'their company')}, "
        f"asking for a referral for a specific role.",
        "",
        "Candidate background:",
        _build_candidate_context(candidate),
        "",
        f"Warmth level: {warmth}",
        f"Tone: {directness}",
        f"Notes on the relationship: {contact.get('notes') or 'No prior contact.'}",
        "",
        "Role being applied for (summary -- do not paste verbatim):",
        jd_text[:800],
        "",
        "Message requirements:",
        "- Reference the specific role title and company; do not paste JD content verbatim",
        "- Make the ask clear but not pressuring",
        "- Keep to 3-4 short paragraphs",
        "- No hollow openers",
    ]

    if warmth_instruction:
        parts.append(f"- {warmth_instruction}")

    if referral_bonus:
        parts.append(
            f"- The referral bonus for this role is {referral_bonus}. Mention this as mutual upside "
            "-- frame it as a benefit to both parties, not as transactional pressure."
        )

    parts.extend([
        "",
        "First, on a single line, write a one-sentence role-fit rationale the candidate can use "
        "to calibrate whether this message is accurate before sending.",
        "Separate it from the message with exactly this delimiter on its own line: ---ROLE-FIT---",
        "Format: [one-line rationale]\\n---ROLE-FIT---\\n[message text]",
    ])

    return "\n".join(parts)


def _build_stage3_prompt(contact: dict, candidate: dict) -> str:
    warmth = contact.get("warmth", "Cold")
    weight = "light and low-pressure" if warmth.lower() == "cold" else "warm and direct"

    return "\n".join([
        f"Write a brief follow-up message from {os.getenv('CANDIDATE_NAME', '[CANDIDATE]')} "
        f"to {contact['contact_name']} at {contact.get('company', 'their company')}.",
        "",
        "Context: a prior outreach message was sent but has not received a response.",
        f"Warmth level: {warmth}",
        f"Notes: {contact.get('notes') or 'No prior contact.'}",
        "",
        "Requirements:",
        f"- Tone: {weight}",
        "- Acknowledge the prior message briefly -- do not repeat the full pitch",
        "- 2-3 sentences maximum",
        "- Leave the door open -- no pressure",
        "- No hollow openers",
    ])


def _build_stage4_prompt(contact: dict, candidate: dict) -> str:
    return "\n".join([
        f"Write a brief closing message from {os.getenv('CANDIDATE_NAME', '[CANDIDATE]')} "
        f"to {contact['contact_name']} at {contact.get('company', 'their company')} "
        "to close the loop on a role that has now resolved.",
        "",
        f"Warmth level: {contact.get('warmth', 'Cold')}",
        f"Notes: {contact.get('notes') or 'No prior contact.'}",
        "",
        "Requirements:",
        "- Share the outcome briefly (the candidate will fill in: offer accepted / withdrawn / not selected)",
        "- Thank the contact genuinely",
        "- Keep the relationship warm regardless of outcome",
        "- 2-3 sentences maximum",
        "- No hollow closers like 'I hope to stay in touch'",
    ])


# ==============================================
# CHARACTER LIMIT ENFORCEMENT
# ==============================================

def _enforce_char_limit(raw: str, warmth: str, client) -> str:
    """For Stage 1: check connection request against tier limit, re-prompt once if over."""
    parts = raw.split("---FOLLOW-UP---")
    connection_request = parts[0].strip()
    follow_up = parts[1].strip() if len(parts) > 1 else ""

    w = warmth.lower()
    limit = 180 if "acquaintance" in w or "former colleague" in w else 300

    if len(connection_request) <= limit:
        return raw

    retry_prompt = (
        f"The connection request you wrote was {len(connection_request)} characters:\n\n"
        f"{connection_request}\n\n"
        f"That exceeds the {limit}-character limit. Rewrite it to fit within {limit} characters "
        f"-- same structure and warmth, tighter language. Then include the follow-up message "
        f"again after '---FOLLOW-UP---'."
    )
    response = client.messages.create(
        model=MODEL,
        max_tokens=512,
        system=SYSTEM_PROMPT,
        messages=[{"role": "user", "content": strip_pii(retry_prompt)}],
    )
    return response.content[0].text


# ==============================================
# GENERATE MESSAGE (importable pure function)
# ==============================================

def generate_message(stage: int, contact: dict, candidate: dict, jd_text: str = None, client=None) -> str:
    """Generate an outreach message for the given stage and contact.
    Pure function -- no file I/O. Injectable client for testing."""
    if client is None:
        client = Anthropic()

    if stage not in (1, 2, 3, 4):
        raise ValueError(f"Invalid stage: {stage}. Must be 1-4.")
    if stage == 2 and not jd_text:
        raise ValueError("jd_text is required for Stage 2 -- pass --role to load the job description.")

    prompt_map = {
        1: lambda: _build_stage1_prompt(contact, candidate),
        2: lambda: _build_stage2_prompt(contact, candidate, jd_text),
        3: lambda: _build_stage3_prompt(contact, candidate),
        4: lambda: _build_stage4_prompt(contact, candidate),
    }

    response = client.messages.create(
        model=MODEL,
        max_tokens=1024,
        system=SYSTEM_PROMPT,
        messages=[{"role": "user", "content": strip_pii(prompt_map[stage]())}],
    )
    raw = response.content[0].text

    if stage == 1:
        raw = _enforce_char_limit(raw, contact.get("warmth", "Cold"), client)

    return raw


def list_contacts(contacts: list) -> None:
    """Print all contacts as a table sorted by stage ascending."""
    sorted_contacts = sorted(contacts, key=lambda c: (c.get("stage") or 0))
    header = (
        f"{'NAME':<25}  {'COMPANY':<22}  {'STG':<6}  {'STATUS':<12}  {'ROLE':<22}"
    )
    print(header)
    print("-" * len(header))
    for c in sorted_contacts:
        print(
            f"{(c.get('contact_name') or ''):<25}  "
            f"{(c.get('company') or ''):<22}  "
            f"{str(c.get('stage') or ''):<6}  "
            f"{(c.get('status') or ''):<12}  "
            f"{(c.get('role_activated') or ''):<22}"
        )


# ==============================================
# CLI HELPERS
# ==============================================

def _warn_if_stage_mismatch(contact: dict, requested_stage: int) -> None:
    """Print a warning if requested stage differs from contact's current xlsx stage."""
    xlsx_stage = contact.get("stage")
    if xlsx_stage is not None and int(xlsx_stage) != requested_stage:
        print(
            f"Warning: contact_pipeline.xlsx shows {contact['contact_name']} at stage {xlsx_stage}, "
            f"but --stage {requested_stage} was requested. Generating Stage {requested_stage} message anyway."
        )


def _format_stage1_output(raw: str, warmth: str) -> str:
    """Parse Stage 1 raw Claude output and format for display."""
    parts = raw.split("---FOLLOW-UP---")
    connection_request = parts[0].strip()
    follow_up = parts[1].strip() if len(parts) > 1 else ""

    w = warmth.lower()
    limit = 180 if "acquaintance" in w or "former colleague" in w else 300
    char_count = len(connection_request)

    if "acquaintance" in w or "former colleague" in w:
        char_line = f"[{char_count} chars generated | ~{limit - char_count} chars remaining for your fill]"
        if char_count > limit:
            char_line += f"  *** OVER {limit}-CHAR TARGET -- trim before sending ***"
    else:
        char_line = f"[{char_count} / {limit} characters]"
        if char_count > limit:
            char_line += "  *** OVER LIMIT -- trim before sending ***"

    output = f"--- Connection Request ---\n{connection_request}\n\n{char_line}"
    if follow_up:
        output += f"\n\n--- Follow-Up (if already connected) ---\n{follow_up}"
    return output


def _format_stage2_output(raw: str) -> str:
    """Parse Stage 2 raw Claude output and format for display."""
    parts = raw.split("---ROLE-FIT---")
    if len(parts) == 2:
        rationale = parts[0].strip()
        message = parts[1].strip()
        return f"Role fit: {rationale}\n\n{message}"
    return raw


def _build_write_back(stage: int, role: str = None) -> dict:
    """Return the dict of fields to write back on confirm for a given stage."""
    today = date.today()
    if stage == 1:
        return {"stage": 2, "first_contact": today}
    if stage == 2:
        updates = {"stage": 3}
        if role:
            updates["role_activated"] = role
        return updates
    if stage == 3:
        return {"stage": 4}
    if stage == 4:
        return {"status": "Closed"}
    return {}


# ==============================================
# MAIN
# ==============================================

def main():
    parser = argparse.ArgumentParser(description="Phase 6 -- Networking and Outreach Support")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--contact", type=str, help="Contact name (partial match)")
    group.add_argument("--list", action="store_true", help="List all contacts")
    parser.add_argument("--stage", type=int, choices=[1, 2, 3, 4], help="Message stage")
    parser.add_argument("--role", type=str, help="Job package role slug (required at Stage 2)")
    args = parser.parse_args()

    contacts = load_contacts(CONTACTS_TRACKER_PATH)

    if args.list:
        list_contacts(contacts)
        return

    if args.stage is None:
        parser.error("--stage is required when using --contact")

    contact = find_contact(contacts, args.contact)
    _warn_if_stage_mismatch(contact, args.stage)

    if args.stage == 2 and not args.role:
        print("Error: --role is required at Stage 2.")
        sys.exit(1)

    candidate = candidate_config.load()
    jd_text = None
    if args.stage == 2:
        jd_path = f"data/job_packages/{args.role}/job_description.txt"
        if not os.path.exists(jd_path):
            print(f"Error: job description not found at {jd_path}")
            sys.exit(1)
        with open(jd_path, encoding="utf-8") as f:
            jd_text = f.read()

    print(f"\n=== Stage {args.stage} | {contact['contact_name']} @ {contact.get('company', '')} | {contact.get('warmth', '')} ===\n")

    raw = generate_message(args.stage, contact, candidate, jd_text=jd_text)

    if args.stage == 1:
        print(_format_stage1_output(raw, contact.get("warmth", "Cold")))
    elif args.stage == 2:
        print(_format_stage2_output(raw))
    else:
        print(raw)

    answer = input("\nDid you send this? (y/n): ").strip().lower()
    if answer == "y":
        updates = _build_write_back(args.stage, role=args.role)
        update_contact(CONTACTS_TRACKER_PATH, contact["contact_name"], updates)
        print(f"Updated {contact['contact_name']} in contact_pipeline.xlsx.")


if __name__ == "__main__":
    main()
